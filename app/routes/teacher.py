from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, after_this_request, abort, current_app, send_from_directory, make_response
from flask_login import current_user, login_required
from functools import wraps
from app import db
from app.models import User, Question, Assignment, AssignmentQuestion, Submission, Section, SectionAssignment, StudentEnrollment
from datetime import datetime, date, timedelta
import json
import os
from openpyxl.utils import get_column_letter
import bleach
from bleach.css_sanitizer import CSSSanitizer
import pymysql
from app.models.schema_import import SchemaImport
from sqlalchemy import create_engine, text, or_, asc, desc
from flask_wtf import CSRFProtect
from app import login_manager, csrf
import secrets
import string
import pandas as pd
from io import StringIO
import logging

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_teacher():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# Configure allowed HTML tags and attributes for rich text editor
ALLOWED_TAGS = [
    'p', 'br', 'strong', 'em', 'u', 's', 'strike', 'ol', 'ul', 'li', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
    'blockquote', 'pre', 'code', 'div', 'span', 'hr', 'table', 'thead', 'tbody', 'tr', 'th', 'td',
    'a', 'img', 'sub', 'sup', 'b', 'i'
]

ALLOWED_ATTRIBUTES = {
    '*': ['class', 'style', 'data-*'],  # Allow style on all elements
    'a': ['href', 'target', 'rel'],
    'img': ['src', 'alt', 'width', 'height', 'data-src'],
    'td': ['colspan', 'rowspan', 'align'],
    'th': ['colspan', 'rowspan', 'align'],
    'div': ['align'],
    'p': ['align'],
    'span': ['data-*'],
    'li': ['class', 'data-list'],
    'ul': ['class'],
    'ol': ['class']
}

# Define allowed CSS styles
ALLOWED_STYLES = [
    'color', 'background-color', 'font-size', 'text-align', 'margin', 'padding',
    'font-weight', 'font-style', 'text-decoration', 'width', 'height', 'display',
    'float', 'text-indent', 'font-family', 'line-height', 'list-style-type',
    # Add specific color values
    '#000000', '#e60000', '#ff9900', '#ffff00', '#008a00', '#0066cc', '#9933ff', '#ffffff',
    '#facccc', '#ffebcc', '#ffffcc', '#cce8cc', '#cce0f5', '#ebd6ff', '#bbbbbb',
    '#f06666', '#ffc266', '#ffff66', '#66b966', '#66a3e0', '#c285ff', '#888888',
    '#a10000', '#b26b00', '#b2b200', '#006100', '#0047b2', '#6b24b2', '#444444',
    '#5c0000', '#663d00', '#666600', '#003700', '#002966', '#3d1466'
]

def sanitize_html(html_content):
    """Sanitize HTML content to prevent XSS attacks"""
    if not html_content:
        return ""
    
    # Create CSS sanitizer with allowed styles
    css_sanitizer = CSSSanitizer(allowed_css_properties=ALLOWED_STYLES)
    
    # Clean the HTML with only allowed tags, attributes, and styles
    clean_html = bleach.clean(
        html_content,
        tags=ALLOWED_TAGS,
        attributes=ALLOWED_ATTRIBUTES,
        css_sanitizer=css_sanitizer,
        strip=True
    )
    return clean_html

teacher = Blueprint('teacher', __name__, url_prefix='/teacher')

@teacher.before_request
def check_teacher():
    if not current_user.is_authenticated:
        flash('Access denied. You must be a teacher to view this page.', 'danger')
        return redirect(url_for('main.index'))
    
    if not current_user.is_active:
        flash('Access denied. Your account has been deactivated. Please contact an administrator.', 'danger')
        return redirect(url_for('main.index'))
    
    if not current_user.is_teacher():
        flash('Access denied. You must be a teacher to view this page.', 'danger')
        return redirect(url_for('main.index'))

@teacher.route('/dashboard')
@login_required
def dashboard():
    assignments = Assignment.query.filter_by(creator_id=current_user.id).all()
    questions = Question.query.filter_by(author_id=current_user.id).all()
    return render_template('teacher/dashboard.html', assignments=assignments, questions=questions)

@teacher.route('/questions')
@login_required
def questions():
    questions = Question.query.filter_by(author_id=current_user.id).all()
    return render_template('teacher/questions.html', questions=questions)

@teacher.route('/question/new', methods=['GET', 'POST'])
@login_required
@teacher_required
def new_question():
    # Get all imported schemas for this teacher
    schemas = SchemaImport.query.filter_by(created_by=current_user.id).order_by(SchemaImport.created_at.desc()).all()
    
    if request.method == 'POST':
        title = request.form.get('title')
        description = sanitize_html(request.form.get('description'))
        question_type = request.form.get('question_type')
        difficulty = request.form.get('difficulty')
        correct_answer = request.form.get('correct_answer')
        db_type = request.form.get('db_type')
        schema_import_id = request.form.get('schema_import_id')
        disable_copy_paste = bool(request.form.get('disable_copy_paste'))
        
        if not all([title, description, question_type, difficulty, correct_answer, db_type]):
            flash('Please fill in all required fields.', 'danger')
            return render_template('teacher/new_question.html', schemas=schemas)
            
        try:
            question = Question(
                title=title,
                description=description,
                question_type=question_type,
                difficulty=int(difficulty),
                correct_answer=correct_answer,
                db_type=db_type,
                author_id=current_user.id,
                disable_copy_paste=disable_copy_paste
            )
            
            if db_type == 'imported_schema':
                if not schema_import_id:
                    flash('Please select an imported schema.', 'danger')
                    return render_template('teacher/new_question.html', schemas=schemas)
                question.schema_import_id = schema_import_id
            elif db_type == 'mysql':
                mysql_db_name = request.form.get('mysql_db_name')
                if not mysql_db_name:
                    flash('Please provide a MySQL database name.', 'danger')
                    return render_template('teacher/new_question.html', schemas=schemas)
                question.mysql_db_name = mysql_db_name
            else:  # sqlite
                sample_db_schema = request.form.get('sample_db_schema')
                if not sample_db_schema:
                    flash('Please provide a database schema.', 'danger')
                    return render_template('teacher/new_question.html', schemas=schemas)
                question.sample_db_schema = sample_db_schema
            
            db.session.add(question)
            db.session.commit()
            
            flash('Question created successfully!', 'success')
            return redirect(url_for('teacher.questions'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating question: {str(e)}', 'danger')
            return render_template('teacher/new_question.html', schemas=schemas)
    
    return render_template('teacher/new_question.html', schemas=schemas)

@teacher.route('/question/<int:question_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_question(question_id):
    question = Question.query.get_or_404(question_id)
    
    if question.author_id != current_user.id:
        flash('You can only edit your own questions.', 'danger')
        return redirect(url_for('teacher.questions'))
    
    # Get all imported schemas for this teacher
    schemas = SchemaImport.query.filter_by(created_by=current_user.id).order_by(SchemaImport.created_at.desc()).all()
    
    if request.method == 'POST':
        try:
            # Print form data for debugging

            # for key, value in request.form.items():

        
            question.title = request.form.get('title')
            question.description = sanitize_html(request.form.get('description'))
            question.question_type = request.form.get('question_type')
            question.difficulty = int(request.form.get('difficulty'))
            question.correct_answer = request.form.get('correct_answer')
            question.disable_copy_paste = bool(request.form.get('disable_copy_paste'))
            
            # Handle database type changes
            db_type = request.form.get('db_type')
            question.db_type = db_type
            
            if db_type == 'imported_schema':
                schema_import_id = request.form.get('schema_import_id')
                if not schema_import_id:
                    error_msg = 'Please select an imported schema.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': False, 'error': error_msg})
                    flash(error_msg, 'danger')
                    return render_template('teacher/edit_question.html', question=question, schemas=schemas)
                question.schema_import_id = schema_import_id
                question.mysql_db_name = None
                question.sample_db_schema = None
            elif db_type == 'mysql':
                mysql_db_name = request.form.get('mysql_db_name')
                
                if not mysql_db_name:
                    error_msg = 'MySQL database name is required when using MySQL.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': False, 'error': error_msg})
                    flash(error_msg, 'danger')
                    return render_template('teacher/edit_question.html', question=question, schemas=schemas)
                
                # Validate MySQL database connection
                try:

                    
                    # Try to connect to the database
                    connection = pymysql.connect(
                        host=os.getenv('MYSQL_HOST', ''),
                        user=os.getenv('MYSQL_USER', ''),
                        password=os.getenv('MYSQL_PASSWORD', ''),
                        port=os.getenv('MYSQL_PORT', 3306),
                        database=mysql_db_name
                    )

                    connection.close()
                    
                    question.mysql_db_name = mysql_db_name
                    question.sample_db_schema = None
                    question.schema_import_id = None
                except Exception as e:
                    error_msg = f'Could not connect to MySQL database: {str(e)}'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': False, 'error': error_msg})
                    flash(error_msg, 'danger')
                    return render_template('teacher/edit_question.html', question=question, schemas=schemas)
            else:  # sqlite
                sample_db_schema = request.form.get('sample_db_schema')
                
                # Validate SQLite schema
                if not sample_db_schema:
                    error_msg = 'SQLite database schema is required when using the in-memory database option.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': False, 'error': error_msg})
                    flash(error_msg, 'danger')
                    return render_template('teacher/edit_question.html', question=question, schemas=schemas)
                
                question.sample_db_schema = sample_db_schema
                question.mysql_db_name = None
                question.schema_import_id = None
            

            db.session.commit()

            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': True, 
                    'message': 'Question updated successfully!'
                })
            
            # Regular form submission
            flash('Question updated successfully!', 'success')
            return redirect(url_for('teacher.questions'))
        except Exception as e:
            db.session.rollback()
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'error': str(e)})
                
            flash(f'Error updating question: {str(e)}', 'danger')
            return render_template('teacher/edit_question.html', question=question, schemas=schemas)
    
    return render_template('teacher/edit_question.html', question=question, schemas=schemas)

@teacher.route('/assignments')
@login_required
def assignments():
    assignments = Assignment.query.filter_by(creator_id=current_user.id).all()
    return render_template('teacher/assignments.html', assignments=assignments)

@teacher.route('/assignment/new', methods=['GET', 'POST'])
@login_required
def new_assignment():
    if request.method == 'POST':
        title = request.form.get('title')
        description = sanitize_html(request.form.get('description'))
        due_date_str = request.form.get('due_date')
        due_time_str = request.form.get('due_time')
        
        # Combine date and time if both are provided
        due_date = None
        if due_date_str:
            if due_time_str:
                # Combine date and time
                due_datetime_str = f"{due_date_str} {due_time_str}"
                due_date = datetime.strptime(due_datetime_str, '%Y-%m-%d %H:%M')
            else:
                # Just date, set time to end of day (23:59:59)
                due_datetime_str = f"{due_date_str} 23:59:59"
                due_date = datetime.strptime(due_datetime_str, '%Y-%m-%d %H:%M:%S')
        
        assignment = Assignment(
            title=title,
            description=description,
            due_date=due_date,
            creator_id=current_user.id
        )
        
        db.session.add(assignment)
        db.session.commit()
        
        # Process questions and scores
        question_ids = request.form.getlist('question_ids')
        question_scores = request.form.getlist('question_scores')
        
        # Default to 10 points if no score is provided
        if not question_scores or len(question_scores) != len(question_ids):
            question_scores = [10] * len(question_ids)
        
        for order, (question_id, score) in enumerate(zip(question_ids, question_scores), 1):
            # Ensure score is an integer and at least 1
            try:
                score_value = int(score)
                if score_value < 1:
                    score_value = 1
            except (ValueError, TypeError):
                score_value = 10  # Default if conversion fails
            
            assignment_question = AssignmentQuestion(
                assignment_id=assignment.id,
                question_id=int(question_id),
                order=order,
                score=score_value
            )
            db.session.add(assignment_question)
        
        db.session.commit()
        
        flash('Assignment created successfully!', 'success')
        return redirect(url_for('teacher.assignments'))
    
    questions = Question.query.filter_by(author_id=current_user.id).all()
    today_date = datetime.now().strftime('%Y-%m-%d')
    return render_template('teacher/new_assignment.html', questions=questions, today_date=today_date)

@teacher.route('/assignment/<int:assignment_id>')
@login_required
def view_assignment(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.creator_id != current_user.id:
        flash('You can only view your own assignments.', 'danger')
        return redirect(url_for('teacher.assignments'))
    
    # Get all questions for this assignment ordered by their order field
    assignment_questions = AssignmentQuestion.query.filter_by(assignment_id=assignment.id).order_by(AssignmentQuestion.order).all()
    questions = [aq.question for aq in assignment_questions]
    
    # Create a mapping of question_id to assignment_question for easy access to scores
    question_scores = {aq.question_id: aq.score for aq in assignment_questions}
    
    # Get all submissions for this assignment
    submissions = Submission.query.filter_by(assignment_id=assignment.id).all()
    
    # Group submissions by student
    submissions_by_student = {}
    for submission in submissions:
        student_id = submission.student_id
        if student_id not in submissions_by_student:
            submissions_by_student[student_id] = []
        submissions_by_student[student_id].append(submission)
    
    # Get student information
    students = {}
    student_scores = {}
    max_score = sum(aq.score for aq in assignment_questions)
    
    for student_id in submissions_by_student:
        student = User.query.get(student_id)
        if student:
            students[student_id] = student
            
            # Calculate student's score for this assignment
            student_score = 0
            student_subs = submissions_by_student[student_id]
            
            # Debug output
            
            # For each correct submission, add the points for that question
            for submission in student_subs:
                if submission.is_correct and submission.question_id in question_scores:
                    question_score = question_scores[submission.question_id]
                    student_score += question_score

                else:
                    logging.debug(f"Submission {submission.id} for student {student_id} is not correct or question score not found.")

            
            student_scores[student_id] = student_score

    
    return render_template(
        'teacher/view_assignment.html',
        assignment=assignment,
        questions=questions,
        assignment_questions=assignment_questions,
        submissions_by_student=submissions_by_student,
        students=students,
        student_scores=student_scores,
        max_score=max_score
    )

@teacher.route('/assignment/<int:assignment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_assignment(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    if assignment.creator_id != current_user.id:
        flash('You can only edit your own assignments.', 'danger')
        return redirect(url_for('teacher.assignments'))
    
    if request.method == 'POST':
        assignment.title = request.form.get('title')
        assignment.description = sanitize_html(request.form.get('description'))
        due_date_str = request.form.get('due_date')
        due_time_str = request.form.get('due_time')
        
        # Combine date and time if both are provided
        if due_date_str:
            if due_time_str:
                # Combine date and time
                due_datetime_str = f"{due_date_str} {due_time_str}"
                assignment.due_date = datetime.strptime(due_datetime_str, '%Y-%m-%d %H:%M')
            else:
                # Just date, set time to end of day (23:59:59)
                due_datetime_str = f"{due_date_str} 23:59:59"
                assignment.due_date = datetime.strptime(due_datetime_str, '%Y-%m-%d %H:%M:%S')
        else:
            assignment.due_date = None
        
        # Delete existing assignment questions
        AssignmentQuestion.query.filter_by(assignment_id=assignment.id).delete()
        
        # Add new assignment questions with scores
        question_ids = request.form.getlist('question_ids')
        question_scores = request.form.getlist('question_scores')
        
        # Default to 10 points if no score is provided
        if not question_scores or len(question_scores) != len(question_ids):
            question_scores = [10] * len(question_ids)
        
        for order, (question_id, score) in enumerate(zip(question_ids, question_scores), 1):
            # Ensure score is an integer and at least 1
            try:
                score_value = int(score)
                if score_value < 1:
                    score_value = 1
            except (ValueError, TypeError):
                score_value = 10  # Default if conversion fails
            
            assignment_question = AssignmentQuestion(
                assignment_id=assignment.id,
                question_id=int(question_id),
                order=order,
                score=score_value
            )
            db.session.add(assignment_question)
        
        db.session.commit()
        
        flash('Assignment updated successfully!', 'success')
        return redirect(url_for('teacher.assignments'))
    
    # Get current questions in the assignment with their scores
    current_questions = AssignmentQuestion.query.filter_by(assignment_id=assignment.id).order_by(AssignmentQuestion.order).all()
    current_question_ids = [aq.question_id for aq in current_questions]
    current_question_scores = [aq.score for aq in current_questions]
    
    all_questions = Question.query.filter_by(author_id=current_user.id).all()
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('teacher/edit_assignment.html', 
                          assignment=assignment, 
                          questions=all_questions, 
                          current_question_ids=current_question_ids,
                          current_question_scores=current_question_scores,
                          today_date=today_date)

@teacher.route('/question/simple/new', methods=['GET'])
@login_required
def simple_new_question():
    """A simplified question creation form for testing"""
    return render_template('teacher/simple_new_question.html')

@teacher.route('/questions/delete/<int:question_id>')
@login_required
def delete_question(question_id):
    question = Question.query.get_or_404(question_id)
    
    # Check ownership
    if question.author_id != current_user.id:
        flash('You can only delete your own questions.', 'danger')
        return redirect(url_for('teacher.questions'))
    
    try:
        # Check if this question is used in any assignments
        assignment_questions = AssignmentQuestion.query.filter_by(question_id=question.id).all()
        if assignment_questions:
            # Get the assignment titles
            assignment_titles = []
            for aq in assignment_questions:
                assignment = Assignment.query.get(aq.assignment_id)
                if assignment:
                    assignment_titles.append(assignment.title)
            
            flash(f'This question cannot be deleted because it is used in the following assignments: {", ".join(assignment_titles)}', 'danger')
            return redirect(url_for('teacher.questions'))
        
        # Check if there are any submissions for this question
        submissions = Submission.query.filter_by(question_id=question.id).count()
        if submissions > 0:
            flash(f'This question cannot be deleted because it has {submissions} student submissions.', 'danger')
            return redirect(url_for('teacher.questions'))
        
        # Store the title for the success message
        title = question.title
        
        # Delete the question
        db.session.delete(question)
        db.session.commit()
        
        flash(f'Question "{title}" has been successfully deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the question: {str(e)}', 'danger')
    
    return redirect(url_for('teacher.questions'))

@teacher.route('/students')
@login_required
def students():
    # Get search parameters
    search_query = request.args.get('search', '')
    sort_field = request.args.get('sort', 'last_name')  # Default sort by last name
    sort_order = request.args.get('order', 'asc')  # Default ascending order
    
    # Get only students enrolled in sections created by this teacher
    teacher_sections = Section.query.filter_by(creator_id=current_user.id).all()
    teacher_section_ids = [section.id for section in teacher_sections]
    
    # Find students from enrollments in these sections
    students_query = db.session.query(User).join(
        StudentEnrollment,
        (StudentEnrollment.student_id == User.id) & (StudentEnrollment.is_active == True)
    ).filter(
        User.role == 'student',
        StudentEnrollment.section_id.in_(teacher_section_ids) if teacher_section_ids else False
    )
    
    # Apply search filter if provided
    if search_query:
        students_query = students_query.filter(
            or_(
                User.first_name.ilike(f'%{search_query}%'),
                User.last_name.ilike(f'%{search_query}%'),
                User.email.ilike(f'%{search_query}%')
            )
        )
    
    # Apply sorting
    if sort_field == 'first_name':
        students_query = students_query.order_by(asc(User.first_name) if sort_order == 'asc' else desc(User.first_name))
    elif sort_field == 'last_name':
        students_query = students_query.order_by(asc(User.last_name) if sort_order == 'asc' else desc(User.last_name))
    elif sort_field == 'email':
        students_query = students_query.order_by(asc(User.email) if sort_order == 'asc' else desc(User.email))
    
    # Group by user ID and get all students
    students_query = students_query.group_by(User.id)
    students = students_query.all()
    
    # Get assignment stats for each student
    stats = {}
    
    for student in students:
        # Get the sections this student is enrolled in
        student_enrollments = StudentEnrollment.query.filter_by(
            student_id=student.id,
            is_active=True
        ).filter(StudentEnrollment.section_id.in_(teacher_section_ids)).all()
        
        student_section_ids = [enrollment.section_id for enrollment in student_enrollments]
        
        # Get assignments assigned to the student's sections
        section_assignments = SectionAssignment.query.filter(
            SectionAssignment.section_id.in_(student_section_ids)
        ).all()
        
        assignment_ids = [sa.assignment_id for sa in section_assignments]
        
        # Get all submissions for this student on their assigned assignments
        submissions = Submission.query.filter_by(student_id=student.id).filter(
            Submission.assignment_id.in_(assignment_ids)
        ).all() if assignment_ids else []
        
        # Get unique assignments that this student has submitted to
        submitted_assignment_ids = set(s.assignment_id for s in submissions)
        
        # Count correct submissions
        correct_submissions = [s for s in submissions if s.is_correct]
        
        stats[student.id] = {
            'total_assignments': len(assignment_ids),
            'started_assignments': len(submitted_assignment_ids),
            'total_submissions': len(submissions),
            'correct_submissions': len(correct_submissions)
        }
    
    return render_template(
        'teacher/students.html', 
        students=students, 
        stats=stats, 
        search_query=search_query,
        sort_field=sort_field,
        sort_order=sort_order
    )

@teacher.route('/student/<int:student_id>')
@login_required
def view_student(student_id):
    # Get the student
    student = User.query.filter_by(id=student_id, role='student').first_or_404()
    
    # Get the student's enrollments
    student_enrollments = StudentEnrollment.query.filter_by(
        student_id=student.id,
        is_active=True
    ).all()
    student_section_ids = [enrollment.section_id for enrollment in student_enrollments]
    
    # Get sections created by this teacher
    teacher_sections = Section.query.filter_by(creator_id=current_user.id).all()
    teacher_section_ids = [section.id for section in teacher_sections]
    
    # Check if the student is in any of the teacher's sections
    is_in_teacher_section = any(section_id in teacher_section_ids for section_id in student_section_ids)
    
    if not is_in_teacher_section:
        flash('You can only view students enrolled in your sections.', 'danger')
        return redirect(url_for('teacher.students'))
    
    # Get current time for checking due dates
    now = datetime.now()
    
    # Get assignments that are assigned to sections where this student is enrolled
    section_assignments = SectionAssignment.query.filter(
        SectionAssignment.section_id.in_(student_section_ids)
    ).all()
    
    # Get the assignment IDs from section assignments
    assignment_ids = [sa.assignment_id for sa in section_assignments]
    
    # Get all assignments created by this teacher that are assigned to the student's sections
    teacher_assignments = Assignment.query.filter(
        Assignment.creator_id == current_user.id,
        Assignment.id.in_(assignment_ids)
    ).all()
    
    # Get all submissions by this student for this teacher's assignments
    student_submissions = {}
    
    for assignment in teacher_assignments:
        # Get all questions for this assignment
        assignment_questions = AssignmentQuestion.query.filter_by(assignment_id=assignment.id).all()
        question_ids = [aq.question_id for aq in assignment_questions]
        
        # Create a map of question_id to assignment_question for score lookup
        question_to_aq = {aq.question_id: aq for aq in assignment_questions}
        
        # Skip if no questions
        if not question_ids:
            continue
        
        # Get submissions for these questions
        submissions = Submission.query.filter_by(
            student_id=student.id,
            assignment_id=assignment.id
        ).all()
        
        # Keep only submissions for questions that are still part of the assignment
        valid_submissions = [s for s in submissions if s.question_id in question_ids]
        
        # For each question, keep only the most recent submission
        latest_submissions_by_question = {}
        for submission in valid_submissions:
            question_id = submission.question_id
            if question_id not in latest_submissions_by_question or submission.submitted_at > latest_submissions_by_question[question_id].submitted_at:
                latest_submissions_by_question[question_id] = submission
        
        # Add score information to each submission
        for submission in valid_submissions:
            if submission.question_id in question_to_aq:
                aq = question_to_aq[submission.question_id]
                submission.question_score = aq.score
                submission.earned_score = aq.score if submission.is_correct else 0
            else:
                submission.question_score = 0
                submission.earned_score = 0
        
        # Get unique submitted question IDs based on latest submissions
        submitted_question_ids = set(latest_submissions_by_question.keys())
        
        # Calculate if assignment is completed - a student must have submitted to all questions
        is_completed = len(submitted_question_ids) == len(question_ids)
        
        # Count correct submissions in the latest submissions
        correct_count = sum(1 for s in latest_submissions_by_question.values() if s.is_correct)
        
        # Calculate total possible score for this assignment
        total_possible_score = sum(aq.score for aq in assignment_questions)
        
        # Calculate earned score based on latest correct submissions
        earned_score = 0
        for qid, submission in latest_submissions_by_question.items():
            if submission.is_correct and qid in question_to_aq:
                earned_score += question_to_aq[qid].score
        
        # Calculate progress
        if len(question_ids) > 0:
            progress = len(submitted_question_ids) / len(question_ids) * 100
        else:
            progress = 0
        
        student_submissions[assignment.id] = {
            'assignment': assignment,
            'submissions': valid_submissions,
            'question_count': len(question_ids),
            'submitted_count': len(submitted_question_ids),
            'correct_count': correct_count,
            'progress': progress,
            'total_possible_score': total_possible_score,
            'earned_score': earned_score,
            'is_completed': is_completed
        }
    
    return render_template(
        'teacher/view_student.html',
        student=student,
        student_submissions=student_submissions,
        now=now
    )

@teacher.route('/sections')
@login_required
def sections():
    # Get all sections created by this teacher
    sections = Section.query.filter_by(creator_id=current_user.id).all()
    
    # Get stats for each section
    section_stats = {}
    
    for section in sections:
        # Count students enrolled in the section using the StudentEnrollment relationship
        student_count = db.session.query(StudentEnrollment).filter_by(section_id=section.id, is_active=True).count()
        
        # Count assignments for the section
        assignment_count = SectionAssignment.query.filter_by(section_id=section.id).count()
        
        section_stats[section.id] = {
            'student_count': student_count,
            'assignment_count': assignment_count
        }
    
    return render_template('teacher/sections.html', sections=sections, section_stats=section_stats)

@teacher.route('/section/new', methods=['GET', 'POST'])
@login_required
def new_section():
    if request.method == 'POST':
        name = request.form.get('name')
        description = sanitize_html(request.form.get('description'))
        database_name = request.form.get('database_name', '').strip()
        
        if not name:
            flash('Section name is required.', 'danger')
            return redirect(url_for('teacher.new_section'))
        
        # Create new section
        section = Section(
            name=name,
            description=description,
            database_name=database_name if database_name else None,
            creator_id=current_user.id
        )
        
        db.session.add(section)
        db.session.commit()
        
        flash('Section created successfully!', 'success')
        return redirect(url_for('teacher.sections'))
    
    return render_template('teacher/new_section.html')

@teacher.route('/section/<int:section_id>')
@login_required
def view_section(section_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only view your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    # Get all students in this section
    students = section.get_enrolled_students(active_only=True)
    
    # Get all assignments for this section
    section_assignments = SectionAssignment.query.filter_by(section_id=section.id).all()
    assignments = [sa.assignment for sa in section_assignments]
    
    # Create a map of assignment_id to section_assignment for the template
    assignment_status = {sa.assignment_id: sa.is_active for sa in section_assignments}
    
    # Get submission stats for each student
    student_stats = {}
    
    for student in students:
        assignment_ids = [sa.assignment_id for sa in section_assignments]
        
        # Skip if no assignments
        if not assignment_ids:
            student_stats[student.id] = {
                'total_submissions': 0,
                'correct_submissions': 0,
                'completed_assignments': 0
            }
            continue
        
        # Get all submissions for this student for section assignments
        submissions = Submission.query.filter_by(student_id=student.id).filter(
            Submission.assignment_id.in_(assignment_ids)
        ).all()
        
        # Get unique assignment IDs with submissions
        submitted_assignment_ids = set(s.assignment_id for s in submissions)
        
        # Count completed assignments
        completed_assignments = 0
        for assignment_id in assignment_ids:
            # Get all questions for this assignment
            assignment_questions = AssignmentQuestion.query.filter_by(assignment_id=assignment_id).all()
            question_ids = [aq.question_id for aq in assignment_questions]
            
            # Skip if no questions
            if not question_ids:
                continue
            
            # Get all submissions for this assignment
            assignment_submissions = [s for s in submissions if s.assignment_id == assignment_id]
            
            # For each question, keep only the most recent submission
            latest_submissions_by_question = {}
            for submission in assignment_submissions:
                question_id = submission.question_id
                if question_id not in latest_submissions_by_question or submission.submitted_at > latest_submissions_by_question[question_id].submitted_at:
                    latest_submissions_by_question[question_id] = submission
            
            # Get unique question IDs based on latest submissions
            submitted_question_ids = set(latest_submissions_by_question.keys())
            
            # If all questions have submissions, count as completed
            if len(submitted_question_ids) == len(question_ids):
                completed_assignments += 1
        
        # Count correct submissions
        correct_submissions = len([s for s in submissions if s.is_correct])
        
        student_stats[student.id] = {
            'total_submissions': len(submissions),
            'correct_submissions': correct_submissions,
            'completed_assignments': completed_assignments
        }
    
    return render_template('teacher/view_section.html', 
                           section=section, 
                           students=students, 
                           assignments=assignments,
                           section_assignments=section_assignments,
                           assignment_status=assignment_status,
                           student_stats=student_stats)

@teacher.route('/section/<int:section_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_section(section_id):
    section = Section.query.get_or_404(section_id)
    
    if section.creator_id != current_user.id:
        flash('You can only edit sections you created.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    if request.method == 'POST':
        section.name = request.form.get('name')
        section.description = sanitize_html(request.form.get('description'))
        database_name = request.form.get('database_name', '').strip()
        section.database_name = database_name if database_name else None
        
        db.session.commit()
        
        flash('Section updated successfully!', 'success')
        return redirect(url_for('teacher.view_section', section_id=section.id))
    
    return render_template('teacher/edit_section.html', section=section)

@teacher.route('/section/<int:section_id>/students', methods=['GET', 'POST'])
@login_required
def manage_section_students(section_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only manage your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    if request.method == 'POST':
        # Get selected student IDs
        student_ids = request.form.getlist('student_ids')
        
        # Verify that these students are either not in any section or in one of this teacher's sections
        if student_ids:
            teacher_sections = Section.query.filter_by(creator_id=current_user.id).all()
            teacher_section_ids = [s.id for s in teacher_sections]
            
            # Query to check if any selected student is in a section not owned by this teacher
            invalid_students_query = db.session.query(User).join(
                StudentEnrollment, StudentEnrollment.student_id == User.id
            ).filter(
                User.id.in_(student_ids),
                StudentEnrollment.section_id.notin_(teacher_section_ids),
                StudentEnrollment.is_active == True
            )
            
            invalid_students = invalid_students_query.all()
            
            if invalid_students:
                invalid_names = [s.username for s in invalid_students]
                flash(f'Cannot add students that belong to other teachers\' sections: {", ".join(invalid_names)}', 'danger')
                return redirect(url_for('teacher.manage_section_students', section_id=section.id))
        
        # First, deactivate all existing enrollments for this section
        StudentEnrollment.query.filter_by(section_id=section.id).update({StudentEnrollment.is_active: False})
        
        # Add or reactivate enrollments for selected students
        if student_ids:
            for student_id in student_ids:
                # Check if enrollment already exists
                enrollment = StudentEnrollment.query.filter_by(
                    student_id=student_id,
                    section_id=section.id
                ).first()
                
                if enrollment:
                    # Reactivate existing enrollment
                    enrollment.is_active = True
                else:
                    # Create new enrollment
                    new_enrollment = StudentEnrollment(
                        student_id=student_id,
                        section_id=section.id,
                        is_active=True
                    )
                    db.session.add(new_enrollment)
        
        db.session.commit()
        
        flash('Students updated successfully!', 'success')
        return redirect(url_for('teacher.view_section', section_id=section.id))
    
    # Get current students in the section
    current_students = section.get_enrolled_students(active_only=True)
    current_student_ids = [s.id for s in current_students]
    
    # Get all sections created by this teacher
    teacher_sections = Section.query.filter_by(creator_id=current_user.id).all()
    teacher_section_ids = [s.id for s in teacher_sections]
    
    # Get all available students - using a simpler approach to avoid auto-correlation issues
    # Get all student users
    all_students = User.query.filter_by(role='student').all()
    
    # Get students with enrollments in sections not owned by this teacher
    students_in_other_sections_query = db.session.query(User.id).join(
        StudentEnrollment, 
        StudentEnrollment.student_id == User.id
    ).filter(
        StudentEnrollment.section_id.notin_(teacher_section_ids) if teacher_section_ids else True,
        StudentEnrollment.is_active == True
    ).distinct()
    
    students_in_other_sections = [r[0] for r in students_in_other_sections_query.all()]
    
    # Filter out students that are in other teachers' sections
    available_students = [s for s in all_students if s.id not in students_in_other_sections]
    
    # Get all students that are in other teachers' sections - simpler approach
    assigned_elsewhere_students = [s for s in all_students if s.id in students_in_other_sections]
    
    return render_template('teacher/manage_section_students.html', 
                           section=section, 
                           all_students=available_students,
                           assigned_elsewhere_students=assigned_elsewhere_students,
                           current_student_ids=current_student_ids)

@teacher.route('/section/<int:section_id>/assignments', methods=['GET', 'POST'])
@login_required
def manage_section_assignments(section_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only manage your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    if request.method == 'POST':
        # Get selected assignment IDs
        assignment_ids = request.form.getlist('assignment_ids')
        
        # First, delete existing section assignments
        SectionAssignment.query.filter_by(section_id=section.id).delete()
        
        # Add new section assignments
        for assignment_id in assignment_ids:
            section_assignment = SectionAssignment(
                section_id=section.id,
                assignment_id=int(assignment_id)
            )
            db.session.add(section_assignment)
        
        db.session.commit()
        
        flash('Assignments updated successfully!', 'success')
        return redirect(url_for('teacher.view_section', section_id=section.id))
    
    # Get current assignments for this section
    current_assignments = SectionAssignment.query.filter_by(section_id=section.id).all()
    current_assignment_ids = [sa.assignment_id for sa in current_assignments]
    
    # Get all assignments created by this teacher
    all_assignments = Assignment.query.filter_by(creator_id=current_user.id).all()
    
    return render_template('teacher/manage_section_assignments.html',
                          section=section,
                          all_assignments=all_assignments,
                          current_assignment_ids=current_assignment_ids)

@teacher.route('/section/<int:section_id>/invitation', methods=['GET', 'POST'])
@login_required
def section_invitation(section_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only manage your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    if request.method == 'POST':
        # Generate new invitation token
        token = section.generate_invitation_token()
        db.session.commit()
        flash('New invitation link generated successfully!', 'success')
    
    # Generate invitation URL if token exists
    invitation_url = None
    if section.invitation_token:
        invitation_url = url_for('auth.register_with_token', 
                                token=section.invitation_token, 
                                _external=True)
    
    return render_template('teacher/section_invitation.html', 
                          section=section, 
                          invitation_url=invitation_url)

@teacher.route('/section/<int:section_id>/invitation/revoke', methods=['POST'])
@login_required
def revoke_invitation(section_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only manage your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    section.invitation_token = None
    db.session.commit()
    
    flash('Invitation link has been revoked.', 'success')
    return redirect(url_for('teacher.section_invitation', section_id=section.id))

@teacher.route('/section/delete/<int:section_id>', methods=['GET'])
@login_required
def delete_section(section_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only delete your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    try:
        # Store the name for the success message
        section_name = section.name
        
        # Delete section
        # Note: The model defines cascade='all, delete-orphan' for enrollments and section_assignments,
        # so they will be automatically deleted
        db.session.delete(section)
        db.session.commit()
        
        flash(f'Section "{section_name}" has been successfully deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the section: {str(e)}', 'danger')
    
    return redirect(url_for('teacher.sections'))

@teacher.route('/section/<int:section_id>/export_results')
@login_required
def export_section_results(section_id):
    """Export section results as Excel file with student scores for all assignments."""
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only export results from your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    # Get all students in this section
    students = section.get_enrolled_students(active_only=True)
    
    # Get all assignments for this section
    section_assignments = SectionAssignment.query.filter_by(section_id=section.id).all()
    assignments = [sa.assignment for sa in section_assignments]
    
    # Sort assignments by ID or creation date
    assignments.sort(key=lambda a: a.id)
    
    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.cell import get_column_letter
    from flask import send_file
    import tempfile
    import os
    
    # Create a workbook with proper properties
    wb = Workbook()
    
    # Set workbook properties
    wb.properties.creator = "SQL Classroom"
    wb.properties.title = f"Section {section.name} Results"
    wb.properties.description = f"Student results for section {section.name}"
    
    # Remove default sheet and create a new one with valid name
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)
    
    # Create worksheet with properly escaped name (limit to 31 chars, no special chars)
    sheet_name = ''.join(c for c in section.name if c.isalnum() or c in ' _-')[:31]
    if not sheet_name:
        sheet_name = "Results"
    ws = wb.create_sheet(title=sheet_name)
    
    # Create header row
    headers = ["No.", "Last Name", "First Name", "Section"]
    
    # For each assignment, add both detailed score and raw score columns
    for assignment in assignments:
        # Limit title length to avoid Excel issues and remove invalid chars
        title = ''.join(c for c in assignment.title if c not in '"*:<>?/\\|')
        if len(title) > 100:
            title = title[:100] + "..."
            
        # headers.append(f"Assignment {assignment.id}: {title}")
        # headers.append(f"Raw Score {assignment.id}")
        headers.append(f"{title}")
        headers.append("Raw Score")

    headers.append("Total Score")
    headers.append("Raw Total")
    
    # Apply header styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Add data rows
    for idx, student in enumerate(students, 1):
        row_num = idx + 1
        student_total_score = 0
        student_total_possible = 0
        
        # Basic student info - ensure values are strings
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=str(student.last_name or ""))
        ws.cell(row=row_num, column=3, value=str(student.first_name or ""))
        ws.cell(row=row_num, column=4, value=str(section.name))
        
        # For each assignment, calculate and add the score
        col_idx = 5
        for assignment in assignments:
            # Find all questions in the assignment
            assignment_questions = AssignmentQuestion.query.filter_by(assignment_id=assignment.id).all()
            question_ids = [aq.question_id for aq in assignment_questions]
            
            # Create score mapping for each question
            question_scores = {aq.question_id: aq.score for aq in assignment_questions}
            
            # Get all submissions for this student and assignment
            submissions = Submission.query.filter_by(
                student_id=student.id,
                assignment_id=assignment.id
            ).all()
            
            # Get the most recent submission for each question
            latest_submissions = {}
            for submission in submissions:
                question_id = submission.question_id
                if question_id not in latest_submissions or submission.submitted_at > latest_submissions[question_id].submitted_at:
                    latest_submissions[question_id] = submission
            
            # Calculate earned points
            earned_points = 0
            for question_id in question_ids:
                if question_id in latest_submissions and latest_submissions[question_id].is_correct:
                    earned_points += question_scores.get(question_id, 0)
            
            # Calculate total possible points for this assignment
            assignment_total_possible = sum(question_scores.values())
            
            # Update student totals
            student_total_score += earned_points
            student_total_possible += assignment_total_possible
            
            # Add the detailed score to the spreadsheet
            if assignment_total_possible > 0:
                percentage = round((earned_points / assignment_total_possible) * 100)
                score_display = f"{earned_points}/{assignment_total_possible} ({percentage}%)"
            else:
                score_display = "N/A"
            
            # Add the score display
            ws.cell(row=row_num, column=col_idx, value=score_display)
            
            # Add the raw score - ensure it's numeric
            ws.cell(row=row_num, column=col_idx + 1, value=float(earned_points))
            
            # Move to next set of columns
            col_idx += 2
        
        # Add total score
        if student_total_possible > 0:
            percentage = round((student_total_score / student_total_possible) * 100)
            ws.cell(row=row_num, column=col_idx, value=f"{student_total_score}/{student_total_possible} ({percentage}%)")
            ws.cell(row=row_num, column=col_idx + 1, value=float(student_total_score))
        else:
            ws.cell(row=row_num, column=col_idx, value="N/A")
            ws.cell(row=row_num, column=col_idx + 1, value=0)
    
    # Adjust column widths
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        if col_idx == 1:  # Number column
            ws.column_dimensions[column_letter].width = 5
        elif col_idx in (2, 3):  # Name columns
            ws.column_dimensions[column_letter].width = 15
        elif col_idx == 4:  # Section column
            ws.column_dimensions[column_letter].width = 20
        elif "Raw Score" in header or header == "Raw Total":  # Raw score columns
            ws.column_dimensions[column_letter].width = 12
        else:  # Assignment and total columns
            ws.column_dimensions[column_letter].width = 25
    
    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    try:
        # Set proper workbook protection options
        wb.security.lockStructure = False
        
        # Save with explicit options
        wb.save(temp_file.name)
        
        # Send the file
        filename = f"{sheet_name}_results.xlsx"
        return_data = send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
        # Schedule file for deletion after request
        @after_this_request
        def remove_file(response):
            try:
                os.unlink(temp_file.name)
            except Exception as e:
                logging.error(f"Error deleting temporary file {temp_file.name}: {e}")
            return response
        
        return return_data
        
    except Exception as e:

        # Clean up temp file in case of error
        try:
            os.unlink(temp_file.name)
        except:
            logging.error(f"Error deleting temporary file {temp_file.name} after error: {e}")
        flash(f'Error generating Excel file. Please try again.', 'danger')
        return redirect(url_for('teacher.view_section', section_id=section.id))

@teacher.route('/upload_image', methods=['POST'])
@login_required
def upload_image():
    """Handle image uploads from the Quill editor"""
    try:
        # Check if the post request has the file part
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part'})
        
        file = request.files['file']
        
        # If user does not select file, browser also
        # submits an empty part without filename
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'})
        
        # Check if it's a valid image
        if file and file.content_type.startswith('image/'):
            # Create uploads directory if it doesn't exist
            upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static', 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Generate a unique filename
            import uuid
            filename = f"{uuid.uuid4()}_{file.filename}"
            file_path = os.path.join(upload_dir, filename)
            
            # Save the file
            file.save(file_path)
            
            # Return the URL to the image
            image_url = url_for('static', filename=f'uploads/{filename}')
            return jsonify({'success': True, 'url': image_url})
        else:
            return jsonify({'success': False, 'error': 'Invalid file type, must be an image'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}) 

@teacher.route('/assignment/delete/<int:assignment_id>', methods=['GET'])
@login_required
def delete_assignment(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # Make sure the teacher owns this assignment
    if assignment.creator_id != current_user.id:
        flash('You can only delete your own assignments.', 'danger')
        return redirect(url_for('teacher.assignments'))
    
    # Delete all section assignments first (due to foreign key constraints)
    SectionAssignment.query.filter_by(assignment_id=assignment.id).delete()
    
    # Delete all assignment questions
    AssignmentQuestion.query.filter_by(assignment_id=assignment.id).delete()
    
    # Delete all submissions for this assignment
    Submission.query.filter_by(assignment_id=assignment.id).delete()
    
    # Finally delete the assignment
    db.session.delete(assignment)
    db.session.commit()
    
    flash('Assignment deleted successfully!', 'success')
    return redirect(url_for('teacher.assignments'))

@teacher.route('/import_schema', methods=['GET', 'POST'])
@login_required
def import_schema():
    if request.method == 'POST':
        # Get database name and prefix from form and environment
        base_name = request.form.get('db_name')
        prefix_type = request.form.get('prefix_type', 'none')  # none, assignment, or template
        description = request.form.get('description', '')
        
        if not base_name:
            flash('Database name is required', 'danger')
            return redirect(url_for('teacher.import_schema'))
            
        # Get schema file
        if 'schema_file' not in request.files:
            flash('No file uploaded', 'danger')
            return redirect(url_for('teacher.import_schema'))
        
        file = request.files['schema_file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('teacher.import_schema'))
        
        if not file.filename.endswith('.sql'):
            flash('Only .sql files are allowed', 'danger')
            return redirect(url_for('teacher.import_schema'))
        
        # Read and parse SQL file
        try:
            schema_content = file.read().decode('utf-8')
            
            # Store the schema in the database
            schema_import = SchemaImport(
                name=base_name,
                description=description,
                schema_content=schema_content,
                created_by=current_user.id,
                is_template=(prefix_type == 'template')
            )
            db.session.add(schema_import)
            db.session.commit()
            
            flash(f'Successfully imported schema: {base_name}', 'success')
            return redirect(url_for('teacher.import_schema'))
                
        except Exception as e:
            flash(f'Error importing schema: {str(e)}', 'danger')
            return redirect(url_for('teacher.import_schema'))
    
    # Get prefix information from environment for the template
    prefix_student = os.getenv('ASSIGNMENTS_DB_PREFIX', 'student_assignment_')
    prefix_template = os.getenv('TEMPLATE_DB_PREFIX', 'template_assignment_')
    
    # Get all imported schemas
    schemas = SchemaImport.query.filter_by(created_by=current_user.id).order_by(SchemaImport.created_at.desc()).all()
    
    return render_template('teacher/import_schema.html',
                         prefix_student=prefix_student,
                         prefix_template=prefix_template,
                         schemas=schemas)

@teacher.route('/schema/<int:schema_id>')
@login_required
@teacher_required
def get_schema(schema_id):
    schema = SchemaImport.query.get_or_404(schema_id)
    if schema.created_by != current_user.id:
        abort(403)
    return jsonify({'content': schema.schema_content})

@teacher.route('/schema/<int:schema_id>/use', methods=['POST'])
@login_required
@teacher_required
def use_schema(schema_id):
    from app.utils import generate_schema_prefix, parse_schema_statements, modify_create_table_statement, get_prefixed_table_name
    
    schema = SchemaImport.query.get_or_404(schema_id)
    if schema.created_by != current_user.id:
        abort(403)
        
    debug_log = []
    debug_log.append(f"Starting schema deployment for schema ID: {schema_id}")
    debug_log.append(f"Schema name: {schema.name}")
    debug_log.append(f"Schema content length: {len(schema.schema_content) if schema.schema_content else 0}")
    
    try:
        # Connect to the sql_classroom database
        connection = pymysql.connect(
            host=os.getenv('MYSQL_HOST', ''),
            user=os.getenv('MYSQL_USER', ''),
            password=os.getenv('MYSQL_PASSWORD', ''),
            port=os.getenv('MYSQL_PORT', 3306),
            database='sql_classroom',
            connect_timeout=30
        )
        
        debug_log.append("Successfully connected to sql_classroom database")
        
        with connection.cursor() as cursor:
            # Generate a unique prefix for table names
            table_prefix = generate_schema_prefix(current_user.id, schema.id)
            debug_log.append(f"Generated table prefix: {table_prefix}")
            
            # First, drop any existing tables with this prefix
            cursor.execute("SHOW TABLES")
            all_tables = [row[0] for row in cursor.fetchall()]
            tables_to_drop = [table for table in all_tables if table.startswith(table_prefix)]
            debug_log.append(f"Found {len(tables_to_drop)} existing tables to drop: {tables_to_drop}")
            
            # Disable foreign key checks to allow dropping tables with foreign key constraints
            if tables_to_drop:
                cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
                debug_log.append("Disabled foreign key checks for table dropping")
            
            for table in tables_to_drop:
                cursor.execute(f"DROP TABLE IF EXISTS `{table}`")
                debug_log.append(f"Dropped table: {table}")
            
            # Re-enable foreign key checks
            if tables_to_drop:
                cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
                debug_log.append("Re-enabled foreign key checks")
            
            # Parse schema statements
            statements = parse_schema_statements(schema.schema_content)
            debug_log.append(f"Parsed {len(statements)} statements from schema content")
            
            created_tables = []
            table_mappings = {}  # Original table name -> prefixed table name
            
            # First pass: Create tables with prefixed names
            create_table_statements = []
            for i, stmt in enumerate(statements):
                debug_log.append(f"Processing statement {i+1}: {stmt[:100]}...")
                
                if stmt.upper().strip().startswith('CREATE TABLE'):
                    create_table_statements.append(stmt)
                    debug_log.append(f"Found CREATE TABLE statement {len(create_table_statements)}")
                    
                    try:
                        # Extract original table name for mapping
                        table_start = stmt.upper().find('TABLE') + 5
                        table_part = stmt[table_start:].strip()
                        original_table = table_part.split()[0].strip('`').rstrip('(').strip('`')
                        debug_log.append(f"Extracted original table name: '{original_table}'")
                        
                        # Create prefixed table name
                        prefixed_table = get_prefixed_table_name(table_prefix, original_table)
                        table_mappings[original_table] = prefixed_table
                        debug_log.append(f"Mapped '{original_table}' -> '{prefixed_table}'")
                        
                        # Modify the CREATE TABLE statement
                        modified_stmt = modify_create_table_statement(stmt, table_prefix)
                        debug_log.append(f"Modified statement preview: {modified_stmt[:150]}...")
                        
                        cursor.execute(modified_stmt)
                        created_tables.append(prefixed_table)
                        debug_log.append(f"Successfully created table: {prefixed_table}")
                        
                    except pymysql.Error as e:
                        error_msg = f"Error creating table from statement {i+1}: {str(e)}"
                        debug_log.append(error_msg)

                        # Continue with other statements
                    except Exception as e:
                        error_msg = f"Unexpected error processing CREATE TABLE {i+1}: {str(e)}"
                        debug_log.append(error_msg)

            
            debug_log.append(f"Total CREATE TABLE statements found: {len(create_table_statements)}")
            debug_log.append(f"Total tables created: {len(created_tables)}")
            debug_log.append(f"Created tables: {created_tables}")
            
            # Second pass: Handle INSERT and other statements with table references
            non_create_statements = 0
            for i, stmt in enumerate(statements):
                if not stmt.upper().strip().startswith('CREATE TABLE'):
                    non_create_statements += 1
                    try:
                        modified_stmt = stmt
                        replacements_made = 0
                        
                        # Replace table names in the statement with prefixed versions
                        for original_table, prefixed_table in table_mappings.items():
                            # Handle both backticked and non-backticked table names properly
                            import re
                            
                            # First replace backticked table names: `tablename`
                            backticked_pattern = r'`' + re.escape(original_table) + r'`'
                            old_stmt = modified_stmt
                            modified_stmt = re.sub(backticked_pattern, f"`{prefixed_table}`", modified_stmt, flags=re.IGNORECASE)
                            if old_stmt != modified_stmt:
                                replacements_made += 1
                            
                            # Then replace non-backticked table names with word boundaries
                            word_boundary_pattern = r'\b' + re.escape(original_table) + r'\b'
                            old_stmt = modified_stmt
                            modified_stmt = re.sub(word_boundary_pattern, f"`{prefixed_table}`", modified_stmt, flags=re.IGNORECASE)
                            if old_stmt != modified_stmt:
                                replacements_made += 1
                        
                        if modified_stmt != stmt:  # Only execute if we made changes
                            cursor.execute(modified_stmt)
                            debug_log.append(f"Executed non-CREATE statement with {replacements_made} table name replacements")
                        elif stmt.strip():  # Execute original if it's not empty and no replacements needed
                            cursor.execute(stmt)
                            debug_log.append(f"Executed original non-CREATE statement (no table references)")
                            
                    except pymysql.Error as e:
                        error_msg = f"Error executing non-CREATE statement {i+1}: {str(e)}"
                        debug_log.append(error_msg)

                        # Continue with other statements
            
            debug_log.append(f"Processed {non_create_statements} non-CREATE statements")
            
            # Grant permissions to sql_student user for each created table (optional)
            # This can be disabled if the database user doesn't have GRANT privileges
            enable_permission_granting = os.getenv('ENABLE_PERMISSION_GRANTING', 'true').lower() == 'true'
            
            if enable_permission_granting:
                permissions_granted = 0
                permission_errors = 0
                for table in created_tables:
                    try:
                        grant_stmt = f"GRANT SELECT ON `sql_classroom`.`{table}` TO 'sql_student'@{os.getenv('APP_DB_NAME', '')}"
                        cursor.execute(grant_stmt)
                        permissions_granted += 1
                        debug_log.append(f"Granted SELECT permission on table: {table}")
                    except pymysql.Error as e:
                        permission_errors += 1
                        error_msg = f"Warning: Could not grant permissions for table {table}: {str(e)}"
                        debug_log.append(error_msg)
                        # Don't print this as an error since it's not critical
                        if permission_errors <= 3:  # Only show first few warnings
                            logging.warning(error_msg)
                
                if permissions_granted > 0:
                    debug_log.append(f"✅ Granted permissions on {permissions_granted} tables")
                if permission_errors > 0:
                    debug_log.append(f"⚠️  Permission warnings for {permission_errors} tables (not critical)")

            else:
                debug_log.append("ℹ️  Permission granting disabled via configuration")
            
            cursor.execute("FLUSH PRIVILEGES")
            connection.commit()
            debug_log.append("Committed transaction and flushed privileges")
            
            # Update the schema record with the table prefix
            schema.active_schema_name = table_prefix
            db.session.commit()
            debug_log.append(f"Updated schema record with active_schema_name: {table_prefix}")
            
            # Final verification - check what tables actually exist
            cursor.execute("SHOW TABLES")
            final_tables = [row[0] for row in cursor.fetchall()]
            final_prefixed_tables = [t for t in final_tables if t.startswith(table_prefix)]
            debug_log.append(f"Final verification: {len(final_prefixed_tables)} tables exist with prefix {table_prefix}")
            debug_log.append(f"Final tables: {final_prefixed_tables}")
            
            return jsonify({
                'success': True, 
                'table_prefix': table_prefix,
                'tables_created': created_tables,
                'tables_verified': final_prefixed_tables,
                'debug_log': debug_log,
                'message': f'Schema successfully deployed to sql_classroom database with prefix: {table_prefix}'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if connection:
            connection.close()

@teacher.route('/schema/<int:schema_id>/delete', methods=['GET'])
@login_required
@teacher_required
def delete_schema(schema_id):
    schema = SchemaImport.query.get_or_404(schema_id)
    
    # Check ownership
    if schema.created_by != current_user.id:
        flash('You can only delete your own imported schemas.', 'danger')
        return redirect(url_for('teacher.import_schema'))
    
    try:
        # Check if this schema is used in any questions
        questions = Question.query.filter_by(schema_import_id=schema.id).all()
        if questions:
            # Get the question titles
            question_titles = [q.title for q in questions]
            flash(f'This schema cannot be deleted because it is used in the following questions: {", ".join(question_titles)}', 'danger')
            return redirect(url_for('teacher.import_schema'))
        
        # Store the name for the success message
        schema_name = schema.name
        
        # Delete the prefixed tables from sql_classroom database if they exist
        tables_deleted = []
        if schema.active_schema_name:
            try:
                connection = pymysql.connect(
                    host=os.getenv('MYSQL_HOST', ''),
                    user=os.getenv('MYSQL_USER', ''),
                    password=os.getenv('MYSQL_PASSWORD', ''),
                    port=os.getenv('MYSQL_PORT', 3306),
                    database='sql_classroom',  # Connect to sql_classroom database
                    connect_timeout=30
                )
                
                with connection.cursor() as cursor:
                    # Get all tables with this schema's prefix
                    cursor.execute("SHOW TABLES")
                    all_tables = [row[0] for row in cursor.fetchall()]
                    tables_to_delete = [table for table in all_tables if table.startswith(schema.active_schema_name)]
                    
                    # Disable foreign key checks to allow dropping tables with foreign key constraints
                    if tables_to_delete:
                        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
                    
                    # Drop each prefixed table
                    for table in tables_to_delete:
                        try:
                            cursor.execute(f"DROP TABLE IF EXISTS `{table}`")
                            tables_deleted.append(table)

                        except Exception as e:
                            logging.error(f"Error dropping table {table}: {e}")
                    # Re-enable foreign key checks
                    if tables_to_delete:
                        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")

                    
                    connection.commit()
                
                connection.close()
                
            except Exception as e:
                # Continue with schema deletion even if table cleanup fails
                logging.error(f"Error cleaning up tables for schema {schema.id}: {e}")
        
        # Delete the schema record from the database
        db.session.delete(schema)
        db.session.commit()
        
        if tables_deleted:
            flash(f'Schema "{schema_name}" and {len(tables_deleted)} associated tables have been successfully deleted.', 'success')
        else:
            flash(f'Schema "{schema_name}" has been successfully deleted.', 'success')
            
    except Exception as e:
        db.session.rollback()
        flash(f'An error occurred while deleting the schema: {str(e)}', 'danger')
    
    return redirect(url_for('teacher.import_schema'))


@teacher.route('/admin/schema-monitor')
@login_required
@teacher_required
def schema_monitor():
    """Monitor all schemas in the system (admin view)"""
    # Only allow admin users to access this view
    if not current_user.role == 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('teacher.dashboard'))
    
    try:
        # Get all schemas
        schemas = SchemaImport.query.all()
        
        # Get table statistics from MySQL
        connection = pymysql.connect(
            host=os.getenv('MYSQL_HOST', ''),
            user=os.getenv('MYSQL_USER', ''),
            password=os.getenv('MYSQL_PASSWORD', ''),
            port=os.getenv('MYSQL_PORT', 3306),
            database='sql_classroom',
            connect_timeout=30
        )
        
        schema_stats = []
        with connection.cursor() as cursor:
            cursor.execute("SHOW TABLES")
            all_tables = [row[0] for row in cursor.fetchall()]
            
            for schema in schemas:
                if schema.active_schema_name:
                    # Count tables for this schema
                    schema_tables = [t for t in all_tables if t.startswith(schema.active_schema_name)]
                    
                    # Get table sizes
                    total_size = 0
                    for table in schema_tables:
                        try:
                            cursor.execute(f"""
                                SELECT ROUND(((data_length + index_length) / 1024 / 1024), 2) AS 'size_mb'
                                FROM information_schema.TABLES 
                                WHERE table_schema='sql_classroom' AND table_name='{table}'
                            """)
                            result = cursor.fetchone()
                            if result and result[0]:
                                total_size += float(result[0])
                        except:
                            logging.error(f"Error getting size for table {table} in schema {schema.name}")
                    
                    schema_stats.append({
                        'schema': schema,
                        'table_count': len(schema_tables),
                        'size_mb': round(total_size, 2),
                        'tables': schema_tables
                    })
        
        connection.close()
        
        return render_template('teacher/schema_monitor.html', schema_stats=schema_stats)
        
    except Exception as e:
        flash(f'Error loading schema statistics: {str(e)}', 'danger')
        return redirect(url_for('teacher.dashboard'))


@teacher.route('/debug/question/<int:question_id>')
@login_required
@teacher_required
def debug_question(question_id):
    """Debug route to check question and schema import configuration"""
    question = Question.query.get_or_404(question_id)
    
    # Check ownership
    if question.author_id != current_user.id:
        flash('You can only debug your own questions.', 'danger')
        return redirect(url_for('teacher.dashboard'))
    
    debug_info = {
        'question': {
            'id': question.id,
            'title': question.title,
            'db_type': question.db_type,
            'schema_import_id': question.schema_import_id,
        }
    }
    
    if question.db_type == 'imported_schema':
        if question.schema_import:
            debug_info['schema_import'] = {
                'id': question.schema_import.id,
                'name': question.schema_import.name,
                'active_schema_name': question.schema_import.active_schema_name,
                'created_by': question.schema_import.created_by,
                'content_length': len(question.schema_import.schema_content) if question.schema_import.schema_content else 0
            }
            
            # Check if tables exist in database
            try:
                import pymysql
                connection = pymysql.connect(
                    host=os.getenv('MYSQL_HOST', ''),
                    user=os.getenv('MYSQL_USER', ''),
                    password=os.getenv('MYSQL_PASSWORD', ''),
                    port=os.getenv('MYSQL_PORT', 3306),
                    database='sql_classroom',
                    connect_timeout=30
                )
                
                with connection.cursor() as cursor:
                    cursor.execute("SHOW TABLES")
                    all_tables = [row[0] for row in cursor.fetchall()]
                    
                    # Find tables with this schema's prefix
                    prefix = question.schema_import.active_schema_name
                    schema_tables = [t for t in all_tables if t.startswith(prefix)] if prefix else []
                    
                    debug_info['database_tables'] = {
                        'total_tables': len(all_tables),
                        'schema_tables': schema_tables,
                        'table_count': len(schema_tables)
                    }
                
                connection.close()
                
            except Exception as e:
                debug_info['database_error'] = str(e)
        else:
            debug_info['schema_import'] = None
    
    return jsonify(debug_info)


@teacher.route('/test-schema-query', methods=['POST'])
@login_required
@csrf.exempt
def test_schema_query():
    """Quick test endpoint to verify schema query modification"""
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    question_id = data.get('question_id')
    test_query = data.get('query', 'SELECT * FROM students LIMIT 5')
    
    if not question_id:
        return jsonify({'error': 'Question ID required'}), 400
    
    question = Question.query.get(question_id)
    if not question:
        return jsonify({'error': 'Question not found'}), 404
    
    if question.author_id != current_user.id:
        return jsonify({'error': 'Access denied'}), 403
    
    debug_info = {
        'question_id': question.id,
        'db_type': question.db_type,
        'schema_import_id': question.schema_import_id,
        'original_query': test_query
    }
    
    if question.db_type == 'imported_schema':
        # Get table prefix
        table_prefix = question.get_table_prefix()
        debug_info['table_prefix'] = table_prefix
        
        if question.schema_import:
            debug_info['schema_import'] = {
                'id': question.schema_import.id,
                'name': question.schema_import.name,
                'active_schema_name': question.schema_import.active_schema_name
            }
            
            if table_prefix:
                # Try to modify the query
                from app.utils import parse_schema_statements
                schema_content = question.schema_import.schema_content
                statements = parse_schema_statements(schema_content)
                table_names = []
                
                for stmt in statements:
                    if stmt.upper().strip().startswith('CREATE TABLE'):
                        table_start = stmt.upper().find('TABLE') + 5
                        table_part = stmt[table_start:].strip()
                        table_name = table_part.split()[0].strip('`').rstrip('(').strip('`')
                        table_names.append(table_name)
                
                debug_info['table_names'] = table_names
                
                # Modify query
                import re
                modified_query = test_query
                replacements = []
                
                for table_name in table_names:
                    pattern = r'\b' + re.escape(table_name) + r'\b'
                    prefixed_table = f"{table_prefix}{table_name}"
                    old_query = modified_query
                    modified_query = re.sub(pattern, prefixed_table, modified_query, flags=re.IGNORECASE)
                    
                    if old_query != modified_query:
                        replacements.append(f"{table_name} -> {prefixed_table}")
                
                debug_info['modified_query'] = modified_query
                debug_info['replacements'] = replacements
                
                # Try to execute the query
                try:
                    import pymysql
                    connection = pymysql.connect(
                        host=os.getenv('MYSQL_HOST', ''),
                        user=os.getenv('MYSQL_USER', ''),
                        password=os.getenv('MYSQL_PASSWORD', ''),
                        port=os.getenv('MYSQL_PORT', 3306),
                        database='sql_classroom',
                        cursorclass=pymysql.cursors.DictCursor
                    )
                    
                    with connection.cursor() as cursor:
                        cursor.execute(modified_query)
                        if cursor.description:
                            results = cursor.fetchall()
                            debug_info['query_results'] = {
                                'columns': [col[0] for col in cursor.description],
                                'row_count': len(results),
                                'sample_rows': results[:3]  # First 3 rows
                            }
                        else:
                            debug_info['query_results'] = 'No results (non-SELECT query)'
                    
                    connection.close()
                    debug_info['execution_success'] = True
                    
                except Exception as e:
                    debug_info['execution_error'] = str(e)
                    debug_info['execution_success'] = False
            else:
                debug_info['error'] = 'No table prefix - schema not deployed'
        else:
            debug_info['error'] = 'No schema import found'
    else:
        debug_info['error'] = f'Not an imported schema (db_type: {question.db_type})'
    
    return jsonify(debug_info)


@teacher.route('/schema-status')
@login_required
@teacher_required  
def schema_status():
    """Show status of all schemas for current teacher"""
    schemas = SchemaImport.query.filter_by(created_by=current_user.id).order_by(SchemaImport.created_at.desc()).all()
    
    schema_info = []
    for schema in schemas:
        info = {
            'id': schema.id,
            'name': schema.name,
            'description': schema.description,
            'active_schema_name': schema.active_schema_name,
            'is_deployed': bool(schema.active_schema_name),
            'created_at': schema.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'table_count': 0,
            'tables': []
        }
        
        # Check actual tables in database if deployed
        if schema.active_schema_name:
            try:
                import pymysql
                connection = pymysql.connect(
                    host=os.getenv('MYSQL_HOST', ''),
                    user=os.getenv('MYSQL_USER', ''),
                    password=os.getenv('MYSQL_PASSWORD', ''),
                    port=os.getenv('MYSQL_PORT', 3306),
                    database='sql_classroom',
                    connect_timeout=30
                )
                
                with connection.cursor() as cursor:
                    cursor.execute("SHOW TABLES")
                    all_tables = [row[0] for row in cursor.fetchall()]
                    
                    # Filter tables with this schema's prefix
                    schema_tables = [t for t in all_tables if t.startswith(schema.active_schema_name)]
                    info['table_count'] = len(schema_tables)
                    info['tables'] = schema_tables
                
                connection.close()
                
            except Exception as e:
                info['database_error'] = str(e)
        
        schema_info.append(info)
    
    return render_template('teacher/schema_status.html', schemas=schema_info)


@teacher.route('/schema/<int:schema_id>/content')
@login_required
@teacher_required
def view_schema_content(schema_id):
    """View the raw content of a schema for debugging"""
    schema = SchemaImport.query.get_or_404(schema_id)
    
    if schema.created_by != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('teacher.import_schema'))
    
    # Parse the schema to show structure
    from app.utils import parse_schema_statements
    statements = parse_schema_statements(schema.schema_content)
    
    parsed_info = {
        'total_statements': len(statements),
        'create_tables': [],
        'other_statements': []
    }
    
    for stmt in statements:
        if stmt.upper().strip().startswith('CREATE TABLE'):
            # Extract table name
            try:
                table_start = stmt.upper().find('TABLE') + 5
                table_part = stmt[table_start:].strip()
                table_name = table_part.split()[0].strip('`').rstrip('(').strip('`')
                parsed_info['create_tables'].append({
                    'name': table_name,
                    'statement': stmt[:200] + '...' if len(stmt) > 200 else stmt
                })
            except Exception as e:
                parsed_info['create_tables'].append({
                    'name': f'ERROR: {str(e)}',
                    'statement': stmt[:200] + '...' if len(stmt) > 200 else stmt
                })
        else:
            parsed_info['other_statements'].append(stmt[:100] + '...' if len(stmt) > 100 else stmt)
    
    return jsonify({
        'schema': {
            'id': schema.id,
            'name': schema.name,
            'active_schema_name': schema.active_schema_name,
            'content_length': len(schema.schema_content)
        },
        'content': schema.schema_content,
        'parsed': parsed_info
    })

@teacher.route('/api/preview-question', methods=['POST'])
@login_required
@csrf.exempt
def preview_question():
    """API endpoint to preview and test question execution for teachers."""
    # Get JSON data from request
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    # Extract data from request
    query = data.get('query', '').strip()
    title = data.get('title', '')
    description = data.get('description', '')
    db_type = data.get('db_type', 'sqlite')
    sample_db_schema = data.get('sample_db_schema', '')
    mysql_db_name = data.get('mysql_db_name', '')
    schema_import_id = data.get('schema_import_id')
    correct_answer = data.get('correct_answer', '')
    check_answer = data.get('check_answer', False)
    
    # Basic validation
    if not query:
        return jsonify({'error': 'Query is required'}), 400

    # Initialize response
    response_data = {
        'query_result': None,
        'answer_check': None
    }
    
    # Create a temporary Question object for preview (not saved to database)
    preview_question = Question(
        title=title,
        description=description,
        db_type=db_type,
        sample_db_schema=sample_db_schema,
        mysql_db_name=mysql_db_name,
        correct_answer=correct_answer,
        author_id=current_user.id
    )
    
    # Set schema import if applicable
    if db_type == 'imported_schema' and schema_import_id:
        try:
            schema_import_id = int(schema_import_id)
            schema_import = SchemaImport.query.get(schema_import_id)
            if schema_import and schema_import.created_by == current_user.id:
                preview_question.schema_import_id = schema_import_id
                preview_question.schema_import = schema_import
        except (ValueError, TypeError) as e:
            return jsonify({'error': f'Invalid schema import ID: {str(e)}'}), 400
    
    try:
        # Validate the query to ensure only DQL operations are allowed
        from app.utils import validate_dql_only_query
        validate_dql_only_query(query)
        
        # Handle different database types
        if db_type == 'mysql' or db_type == 'imported_schema':
            try:
                # Execute the query using a similar but simplified approach to student's route
                if db_type == 'mysql':
                    # Validate MySQL database connection
                    if not mysql_db_name:
                        return jsonify({'error': 'MySQL database name is required'}), 400
                    
                    # Try to connect to the database
                    connection = pymysql.connect(
                        host=os.getenv('MYSQL_HOST', ''),
                        user=os.getenv('MYSQL_USER', ''),
                        password=os.getenv('MYSQL_PASSWORD', ''),
                        port=os.getenv('MYSQL_PORT', 3306),
                        database=mysql_db_name,
                        cursorclass=pymysql.cursors.DictCursor
                    )
                else:  # imported_schema
                    if not preview_question.schema_import:
                        return jsonify({'error': 'Invalid or missing schema import. Please ensure the schema is properly imported and deployed.'}), 400
                    
                    if not preview_question.schema_import.active_schema_name:
                        return jsonify({'error': 'Schema import exists but is not deployed. Please use the "Use" button to deploy the schema first.'}), 400
                    
                    # Get the actual database name from the schema import
                    imported_schema_name = preview_question.schema_import.active_schema_name
                    
                    # Connect to SQL classroom database first
                    connection = pymysql.connect(
                        host=os.getenv('MYSQL_HOST', ''),
                        user=os.getenv('MYSQL_USER', ''),
                        password=os.getenv('MYSQL_PASSWORD', ''),
                        port=os.getenv('MYSQL_PORT', 3306),
                        database='sql_classroom',
                        cursorclass=pymysql.cursors.DictCursor
                    )
                    
                    # Modify query to use table prefixes if needed
                    table_prefix = preview_question.get_table_prefix()

                    
                    if table_prefix:
                        # Get all table names from the schema content
                        schema_content = preview_question.schema_import.schema_content

                        
                        table_names = []
                        
                        # Parse CREATE TABLE statements to extract table names
                        from app.utils import parse_schema_statements
                        statements = parse_schema_statements(schema_content)
                        
                        for i, stmt in enumerate(statements):

                            if stmt.upper().strip().startswith('CREATE TABLE'):

                                # Extract table name using the same logic as import
                                table_start = stmt.upper().find('TABLE') + 5
                                table_part = stmt[table_start:].strip()

                                table_name = table_part.split()[0].strip('`').rstrip('(').strip('`')

                                table_names.append(table_name)
                        


                        
                        # Replace table names with prefixed versions in the query using word boundaries
                        import re
                        modified_query = query
                        for table_name in table_names:
                            # Use word boundaries to avoid partial matches
                            pattern = r'\b' + re.escape(table_name) + r'\b'
                            prefixed_table = f"{table_prefix}{table_name}"
                            old_query = modified_query
                            modified_query = re.sub(pattern, prefixed_table, modified_query, flags=re.IGNORECASE)
                            if old_query != modified_query:
                                logging.debug(f"Table name '{table_name}' replaced with '{prefixed_table}' in query.")
                            else:
                                logging.debug(f"No replacement made for table name '{table_name}'.")

                        query = modified_query

                    else:
                        logging.warning("No table prefix found for imported schema. Query may not execute as expected.")
                
                # Execute query and fetch results
                with connection.cursor() as cursor:
                    # Execute the query
                    cursor.execute(query)
                    
                    # Get results
                    if cursor.description:
                        data = cursor.fetchall()
                        
                        # Get column names from cursor description
                        columns = [col[0] for col in cursor.description]
                        
                        # Convert dictionary data to lists for JSON response
                        rows = []
                        for row in data:
                            # Convert any non-JSON serializable objects to strings
                            serializable_row = {}
                            for key, value in row.items():
                                if isinstance(value, (dict, list, str, int, float, bool, type(None))):
                                    serializable_row[key] = value
                                else:
                                    serializable_row[key] = str(value)
                            
                            # Ensure values are in the same order as columns
                            row_values = [serializable_row.get(col) for col in columns]
                            rows.append(row_values)
                    else:
                        # For non-SELECT queries
                        columns = ['Result']
                        rows = [['Query executed successfully, but returned no rows.']]
                
                connection.close()
                
                # Set query result in response
                response_data['query_result'] = {
                    'columns': columns,
                    'data': rows
                }
                
            except pymysql.err.OperationalError as e:
                error_code, error_message = e.args
                return jsonify({'error': f'Database error: {error_message}'}), 400
            except pymysql.err.ProgrammingError as e:
                error_code, error_message = e.args
                return jsonify({'error': f'SQL error: {error_message}'}), 400
            except Exception as e:
                return jsonify({'error': str(e)}), 400
                
        elif db_type == 'sqlite':
            # Execute SQLite in-memory database query
            if not sample_db_schema:
                return jsonify({'error': 'SQLite schema is required'}), 400
                
            try:
                # Create in-memory SQLite database
                engine = create_engine('sqlite:///:memory:', 
                                     connect_args={'check_same_thread': False})
                
                # Create a new connection and load the schema
                with engine.connect() as conn:
                    # Load schema
                    schema_statements = sample_db_schema.split(';')
                    trans = conn.begin()
                    try:
                        for statement in schema_statements:
                            if statement.strip():
                                conn.execute(text(statement))
                        trans.commit()
                    except Exception as e:
                        trans.rollback()
                        return jsonify({'error': f'Error in schema: {str(e)}'}), 400
                
                    # Execute the query
                    result_proxy = conn.execute(text(query))
                    
                    # Get columns and data
                    columns = list(result_proxy.keys())
                    data = [list(row) for row in result_proxy]
                    
                    # Set query result in response
                    response_data['query_result'] = {
                        'columns': columns,
                        'data': data
                    }
                    
            except Exception as e:
                return jsonify({'error': f'Error executing query: {str(e)}'}), 400
        else:
            return jsonify({'error': f'Unsupported database type: {db_type}'}), 400
            
        # If checking against correct answer
        if check_answer and correct_answer:
            # Create a simple check - just compare query results
            correct_result = None
            
            # Execute the correct answer query and get its results
            try:
                if db_type == 'mysql' or db_type == 'imported_schema':
                    # Re-establish connection
                    if db_type == 'mysql':
                        connection = pymysql.connect(
                            host=os.getenv('MYSQL_HOST', ''),
                            user=os.getenv('MYSQL_USER', ''),
                            password=os.getenv('MYSQL_PASSWORD', ''),
                            port=os.getenv('MYSQL_PORT', 3306),
                            database=mysql_db_name,
                            cursorclass=pymysql.cursors.DictCursor
                        )
                    else:
                        connection = pymysql.connect(
                            host=os.getenv('MYSQL_HOST', ''),
                            user=os.getenv('MYSQL_USER', ''),
                            password=os.getenv('MYSQL_PASSWORD', ''),
                            port=os.getenv('MYSQL_PORT', 3306),
                            database='sql_classroom',
                            cursorclass=pymysql.cursors.DictCursor
                        )
                        
                        # Apply table prefixes to correct answer if needed
                        if table_prefix:
                            modified_answer = correct_answer
                            for table_name in table_names:
                                modified_answer = modified_answer.replace(table_name, f"{table_prefix}{table_name}")
                            correct_answer = modified_answer
                    
                    with connection.cursor() as cursor:
                        cursor.execute(correct_answer)
                        correct_data = cursor.fetchall()
                        correct_columns = [col[0] for col in cursor.description]
                        
                        # Convert to list format for comparison
                        correct_rows = []
                        for row in correct_data:
                            # Normalize data for comparison
                            row_values = [row[col] if not isinstance(row[col], (datetime, date)) 
                                         else str(row[col]) for col in correct_columns]
                            correct_rows.append(row_values)
                            
                    connection.close()
                    correct_result = {
                        'columns': correct_columns,
                        'data': correct_rows
                    }
                    
                elif db_type == 'sqlite':
                    # Create a new in-memory database for correct answer
                    correct_engine = create_engine('sqlite:///:memory:',
                                                  connect_args={'check_same_thread': False})
                    
                    with correct_engine.connect() as conn:
                        # Load schema again
                        schema_statements = sample_db_schema.split(';')
                        trans = conn.begin()
                        for statement in schema_statements:
                            if statement.strip():
                                conn.execute(text(statement))
                        trans.commit()
                        
                        # Execute correct answer
                        correct_result_proxy = conn.execute(text(correct_answer))
                        correct_columns = list(correct_result_proxy.keys())
                        correct_data = [list(row) for row in correct_result_proxy]
                        
                        correct_result = {
                            'columns': correct_columns,
                            'data': correct_data
                        }
                
                # Compare results
                user_data = response_data['query_result']['data']
                correct_data = correct_result['data']
                
                # Normalize and compare row data
                is_correct = False

                def normalize_value(value):
                    """Convert string representations of numbers back to numeric types for comparison"""
                    if value is None:
                        return None
                    
                    if isinstance(value, str):
                        # Try to convert string to numeric types
                        try:
                            # Check if it's an integer
                            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                                return int(value)
                            
                            # Check if it's a float/decimal
                            try:
                                float_val = float(value)
                                # Check if it's a string representation of a decimal
                                if '.' in value:
                                    # Keep it as float for comparison
                                    return float_val
                                return float_val
                            except ValueError:
                                logging.debug(f"Value '{value}' is not a float or int, keeping as string.")
                        except:
                            # If conversion fails, keep as string
                            logging.debug(f"Value '{value}' could not be converted, keeping as string.")
                    
                    return value

                def normalize_row(row):
                    """Normalize all values in a row for proper comparison"""
                    return [normalize_value(cell) for cell in row]

                def rows_equal(row1, row2):
                    """Compare rows accounting for numeric equality"""
                    if len(row1) != len(row2):
                        return False
                    
                    for i in range(len(row1)):
                        # Special comparison for floats - use approximate equality
                        if isinstance(row1[i], float) or isinstance(row2[i], float):
                            try:
                                # Convert to float if possible
                                val1 = float(row1[i]) if row1[i] is not None else None
                                val2 = float(row2[i]) if row2[i] is not None else None
                                
                                # Check for None values
                                if val1 is None or val2 is None:
                                    if val1 != val2:
                                        return False
                                    continue
                                
                                # Use approximate equality for floats (handles precision differences)
                                if abs(val1 - val2) > 0.0001:  # Small epsilon for float comparison
                                    return False
                            except (ValueError, TypeError):
                                # If conversion fails, fall back to direct comparison
                                if row1[i] != row2[i]:
                                    return False
                        else:
                            # For non-float values, use direct comparison
                            if row1[i] != row2[i]:
                                return False
                    
                    return True

                if len(user_data) == len(correct_data):
                    if len(user_data) == 0:
                        # Both returned empty result sets
                        is_correct = True
                    else:
                        # Check if rows match (order matters for simplicity)
                        is_correct = True
                        
                        # Normalize the data first
                        normalized_user_data = [normalize_row(row) for row in user_data]
                        normalized_correct_data = [normalize_row(row) for row in correct_data]
                        
                        for i in range(len(normalized_user_data)):
                            if not rows_equal(normalized_user_data[i], normalized_correct_data[i]):
                                is_correct = False
                                break

                feedback = "Your answer is correct!" if is_correct else "Your answer is incorrect. The expected results do not match your query output."
                
                # Add answer check to response
                response_data['answer_check'] = {
                    'is_correct': is_correct,
                    'feedback': feedback
                }
                
            except Exception as e:
                # Add error feedback to response
                response_data['answer_check'] = {
                    'is_correct': False,
                    'feedback': f"Error checking answer: {str(e)}"
                }
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

@teacher.route('/section/<int:section_id>/assignment/<int:assignment_id>/toggle_status', methods=['POST'])
@login_required
def toggle_section_assignment_status(section_id, assignment_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only manage your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    # Get the section assignment
    section_assignment = SectionAssignment.query.filter_by(
        section_id=section_id, 
        assignment_id=assignment_id
    ).first_or_404()
    
    # Toggle the status
    section_assignment.is_active = not section_assignment.is_active
    
    status = "enabled" if section_assignment.is_active else "disabled"
    
    try:
        db.session.commit()
        flash(f'Assignment has been {status}!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating assignment status: {str(e)}', 'danger')
    
    return redirect(url_for('teacher.view_section', section_id=section_id))

@teacher.route('/section/<int:section_id>/assignment/<int:assignment_id>/export_results')
@login_required
def export_assignment_results(section_id, assignment_id):
    """Export results for a specific assignment in a section as Excel file."""
    section = Section.query.get_or_404(section_id)
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only export results from your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    # Check if the assignment belongs to this section
    section_assignment = SectionAssignment.query.filter_by(
        section_id=section.id, 
        assignment_id=assignment.id
    ).first_or_404()
    
    # Get all students in this section
    students = section.get_enrolled_students(active_only=True)
    
    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.cell import get_column_letter
    from flask import send_file
    import tempfile
    import os
    
    # Create a workbook with proper properties
    wb = Workbook()
    
    # Set workbook properties
    wb.properties.creator = "SQL Classroom"
    wb.properties.title = f"Assignment {assignment.title} Results"
    wb.properties.description = f"Student results for {assignment.title} in section {section.name}"
    
    # Remove default sheet and create a new one with valid name
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)
    
    # Create worksheet with properly escaped name (limit to 31 chars, no special chars)
    assignment_name = ''.join(c for c in assignment.title if c.isalnum() or c in ' _-')[:20]
    if not assignment_name:
        assignment_name = f"Assignment_{assignment.id}"
    sheet_name = assignment_name
    ws = wb.create_sheet(title=sheet_name)
    
    # Find all questions in the assignment
    assignment_questions = AssignmentQuestion.query.filter_by(assignment_id=assignment.id).order_by(AssignmentQuestion.order).all()
    questions = [aq.question for aq in assignment_questions]
    
    # Create header row
    headers = ["No.", "Last Name", "First Name", "Username", "Email"]
    
    # Add a column for each question
    for i, question in enumerate(questions, 1):
        # Limit title length to avoid Excel issues and remove invalid chars
        title = ''.join(c for c in question.title if c not in '"*:<>?/\\|')
        if len(title) > 30:
            title = title[:27] + "..."
        headers.append(f"Q{i}: {title}")
    
    # Add total score
    headers.append("Total Score")
    headers.append("Percentage")
    
    # Apply header styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Calculate total possible points for this assignment
    question_scores = {aq.question_id: aq.score for aq in assignment_questions}
    total_possible = sum(question_scores.values())
    
    # Add data rows
    for idx, student in enumerate(students, 1):
        row_num = idx + 1
        
        # Basic student info
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=str(student.last_name or ""))
        ws.cell(row=row_num, column=3, value=str(student.first_name or ""))
        ws.cell(row=row_num, column=4, value=student.username)
        ws.cell(row=row_num, column=5, value=student.email)
        
        # Get all submissions for this student and assignment
        submissions = Submission.query.filter_by(
            student_id=student.id,
            assignment_id=assignment.id
        ).all()
        
        # Get the most recent submission for each question
        latest_submissions = {}
        for submission in submissions:
            question_id = submission.question_id
            if question_id not in latest_submissions or submission.submitted_at > latest_submissions[question_id].submitted_at:
                latest_submissions[question_id] = submission
        
        # Fill in question results
        col_idx = 6  # Start after student info columns
        earned_points = 0
        
        for question in questions:
            question_id = question.id
            question_score = question_scores.get(question_id, 0)
            
            if question_id in latest_submissions:
                submission = latest_submissions[question_id]
                if submission.is_correct:
                    ws.cell(row=row_num, column=col_idx, value="Correct")
                    earned_points += question_score
                    # Add green fill for correct answers
                    ws.cell(row=row_num, column=col_idx).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws.cell(row=row_num, column=col_idx, value="Incorrect")
                    # Add light red fill for incorrect answers
                    ws.cell(row=row_num, column=col_idx).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                ws.cell(row=row_num, column=col_idx, value="Not Attempted")
            
            col_idx += 1
        
        # Add total score
        ws.cell(row=row_num, column=col_idx, value=f"{earned_points}/{total_possible}")
        
        # Add percentage
        if total_possible > 0:
            percentage = round((earned_points / total_possible) * 100)
            ws.cell(row=row_num, column=col_idx + 1, value=f"{percentage}%")
        else:
            ws.cell(row=row_num, column=col_idx + 1, value="N/A")
    
    # Adjust column widths
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        if col_idx == 1:  # Number column
            ws.column_dimensions[column_letter].width = 5
        elif col_idx in (2, 3):  # Name columns
            ws.column_dimensions[column_letter].width = 15
        elif col_idx == 4:  # Username column
            ws.column_dimensions[column_letter].width = 15
        elif col_idx == 5:  # Email column
            ws.column_dimensions[column_letter].width = 25
        elif "Q" in header:  # Question columns
            ws.column_dimensions[column_letter].width = 20
        else:  # Total columns
            ws.column_dimensions[column_letter].width = 12
    
    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    try:
        # Set proper workbook protection options
        wb.security.lockStructure = False
        
        # Save with explicit options
        wb.save(temp_file.name)
        
        # Send the file
        filename = f"{section.name}_{assignment_name}_results.xlsx"
        safe_filename = ''.join(c for c in filename if c.isalnum() or c in ' _-.()[]{}')
        
        return_data = send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=safe_filename
        )
        
        # Schedule file for deletion after request
        @after_this_request
        def remove_file(response):
            try:
                os.unlink(temp_file.name)
            except Exception as e:
                logging.error(f"Error deleting temporary file {temp_file.name}: {e}")
            return response
        
        return return_data
        
    except Exception as e:

        # Clean up temp file in case of error
        try:
            os.unlink(temp_file.name)
        except:
            logging.error(f"Error deleting temporary file {temp_file.name} after error: {e}")
        flash(f'Error generating Excel file: {str(e)}', 'danger')
        return redirect(url_for('teacher.view_section', section_id=section.id))

@teacher.route('/section/<int:section_id>/assignment/<int:assignment_id>/duplicate', methods=['GET'])
@login_required
def duplicate_section_assignment(section_id, assignment_id):
    section = Section.query.get_or_404(section_id)
    
    # Make sure the teacher owns this section
    if section.creator_id != current_user.id:
        flash('You can only manage your own sections.', 'danger')
        return redirect(url_for('teacher.sections'))
    
    # Get the assignment to duplicate
    original_assignment = Assignment.query.get_or_404(assignment_id)
    
    # Create a new assignment as a copy of the original
    new_assignment = Assignment(
        title=f"{original_assignment.title} (Copy)",
        description=original_assignment.description,
        due_date=original_assignment.due_date,
        creator_id=current_user.id
    )
    
    db.session.add(new_assignment)
    db.session.flush()  # Generate ID for the new assignment
    
    # Copy all questions and their scores from the original assignment
    original_questions = AssignmentQuestion.query.filter_by(assignment_id=original_assignment.id).all()
    
    for aq in original_questions:
        new_aq = AssignmentQuestion(
            assignment_id=new_assignment.id,
            question_id=aq.question_id,
            order=aq.order,
            score=aq.score
        )
        db.session.add(new_aq)
    
    # Add the new assignment to the current section
    section_assignment = SectionAssignment(
        section_id=section.id,
        assignment_id=new_assignment.id,
        is_active=True
    )
    db.session.add(section_assignment)
    
    try:
        db.session.commit()
        flash(f'Assignment "{original_assignment.title}" has been duplicated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error duplicating assignment: {str(e)}', 'danger')
    
    return redirect(url_for('teacher.view_section', section_id=section.id))

@teacher.route('/submission/<int:submission_id>/delete', methods=['POST'])
@login_required
@teacher_required
@csrf.exempt
def delete_submission(submission_id):
    """Delete a student's submission so they can re-answer the question."""
    submission = Submission.query.get_or_404(submission_id)
    
    # Get the assignment to verify teacher ownership
    assignment = Assignment.query.get_or_404(submission.assignment_id)
    
    # Make sure the teacher owns this assignment
    if assignment.creator_id != current_user.id:
        flash('You can only manage submissions for your own assignments.', 'danger')
        return redirect(url_for('teacher.assignments'))
    
    # Get student and question info for the flash message
    student = User.query.get(submission.student_id)
    question = Question.query.get(submission.question_id)
    
    try:
        # Delete the submission
        db.session.delete(submission)
        db.session.commit()
        
        flash(f'Submission deleted successfully! {student.username if student else "Student"} can now re-answer "{question.title if question else "the question"}".', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting submission: {str(e)}', 'danger')
    
    # Return to the assignment view
    return redirect(url_for('teacher.view_assignment', assignment_id=assignment.id))

@teacher.route('/student/<int:student_id>/reset_password', methods=['POST'])
@login_required
@teacher_required
def reset_student_password(student_id):
    # Get the student
    student = User.query.filter_by(id=student_id, role='student').first_or_404()
    
    # Get the student's enrollments
    student_enrollments = StudentEnrollment.query.filter_by(
        student_id=student.id,
        is_active=True
    ).all()
    student_section_ids = [enrollment.section_id for enrollment in student_enrollments]
    
    # Get sections created by this teacher
    teacher_sections = Section.query.filter_by(creator_id=current_user.id).all()
    teacher_section_ids = [section.id for section in teacher_sections]
    
    # Check if the student is in any of the teacher's sections
    is_in_teacher_section = any(section_id in teacher_section_ids for section_id in student_section_ids)
    
    if not is_in_teacher_section:
        flash('You can only reset passwords for students enrolled in your sections.', 'danger')
        return redirect(url_for('teacher.students'))
    
    try:
        # Generate a temporary password
        import secrets
        import string
        
        # Generate a random password with 10 characters
        temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))
        
        # Set the new password for the student
        student.set_password(temp_password)
        db.session.commit()
        
        # Show the temporary password to the teacher
        flash(f'Password for {student.full_name} has been reset. Temporary password: {temp_password}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error resetting password: {str(e)}', 'danger')
    
    # Return to the student details page
    return redirect(url_for('teacher.view_student', student_id=student.id))

@teacher.route('/sql-playground', methods=['GET'])
@login_required
@teacher_required
def sql_playground():
    """Display the SQL Playground page where teachers can practice SQL queries."""
    return render_template('teacher/sql_playground.html')

@teacher.route('/api/playground-execute', methods=['POST'])
@login_required
@teacher_required
@csrf.exempt
def playground_execute():
    """API endpoint to execute SQL queries in the playground."""
    data = request.json
    if not data or 'query' not in data or 'database_name' not in data:
        return jsonify({'error': 'Missing required parameters'}), 400
    
    query = data['query'].strip()
    database_name = data['database_name'].strip()
    
    try:
        # Validate the query to ensure only DQL operations are allowed
        from app.utils import validate_dql_only_query
        validate_dql_only_query(query)
        
        # Check if database access is allowed
        from app.models import AllowedDatabase
        from app.models.schema_import import SchemaImport
        
        # Check if database is in allowed list
        is_allowed_database = AllowedDatabase.is_database_allowed(database_name)
        
        # Check if this is a teacher's own imported schema name
        is_teacher_schema_access = False
        selected_schema = None
        actual_database_name = database_name  # Will be the actual database to connect to
        
        if not is_allowed_database:
            # Check if the database_name is actually a schema name created by this teacher
            teacher_schema = SchemaImport.query.filter_by(
                created_by=current_user.id,
                name=database_name
            ).filter(SchemaImport.active_schema_name.isnot(None)).first()
            
            if teacher_schema:
                is_teacher_schema_access = True
                selected_schema = teacher_schema
                actual_database_name = 'sql_classroom'  # Connect to sql_classroom for schema access
        
        # Allow access if it's an allowed database OR teacher's own schema
        if not (is_allowed_database or is_teacher_schema_access):
            # Get list of allowed databases for error message
            allowed_databases = AllowedDatabase.get_active_database_names()
            error_msg = f'Access to database "{database_name}" is not allowed.'
            
            if allowed_databases:
                db_list = ', '.join(allowed_databases)
                # Also mention schema names if teacher has schemas
                try:
                    teacher_schemas = SchemaImport.query.filter_by(created_by=current_user.id).filter(
                        SchemaImport.active_schema_name.isnot(None)
                    ).all()
                    if teacher_schemas:
                        schema_names = [schema.name for schema in teacher_schemas]
                        db_list += ', ' + ', '.join(schema_names) + ' (your schemas)'
                except:
                    logging.error(f"Error retrieving schemas for teacher {current_user.id}: {e}")

                error_msg += f' Available options: {db_list}'
                
            return jsonify({'error': error_msg}), 403
        
        # For imported schema access, check if we need to rewrite table names
        if selected_schema:
            # Use the selected schema for query rewriting
            table_prefix = selected_schema.active_schema_name
            schema_content = selected_schema.schema_content
            

            
            # Apply query rewriting using utility function
            from app.utils import rewrite_query_for_schema
            query = rewrite_query_for_schema(query, schema_content, table_prefix)
        
        # Teachers now have the same query restrictions as students for security
        try:
            # Connect to MySQL using the actual database name (may be sql_classroom for schema access)
            conn = pymysql.connect(
                host=os.getenv('MYSQL_HOST', ''),
                user=os.getenv('MYSQL_USER', ''),
                password=os.getenv('MYSQL_PASSWORD', ''),
                port=os.getenv('MYSQL_PORT', 3306),
                database=actual_database_name,
                cursorclass=pymysql.cursors.DictCursor
            )
            
            # Execute query and fetch results
            with conn.cursor() as cursor:
                # Execute the query
                cursor.execute(query)
                
                # Check if this is a SELECT query by looking for a result set
                if cursor.description:
                    # For SELECT queries, get the results
                    data = cursor.fetchall()
                    
                    # Get column names directly from cursor description
                    columns = [col[0] for col in cursor.description]
                    
                    # Convert dictionary data to lists for JSON response
                    rows = []
                    for row in data:
                        # Convert any non-JSON serializable objects to strings
                        serializable_row = {}
                        for key, value in row.items():
                            if isinstance(value, (dict, list, str, int, float, bool, type(None))):
                                serializable_row[key] = value
                            else:
                                serializable_row[key] = str(value)
                        
                        # Ensure values are in the same order as columns
                        row_values = [serializable_row.get(col) for col in columns]
                        rows.append(row_values)
                    
                    # Check if this is a SHOW TABLES query on an imported schema
                    from app.utils import is_show_tables_query, process_show_tables_result_for_schema, is_show_databases_query, filter_show_databases_result_for_user
                    if is_show_tables_query(query):
                        if selected_schema:
                            # Filter tables for the specific selected schema
                            prefix = selected_schema.active_schema_name
                            # Filter tables to only show those belonging to this schema
                            filtered_rows = []
                            for row in rows:
                                table_name = row[0]
                                if table_name.startswith(prefix):
                                    # Remove the prefix to show the original table name
                                    original_name = table_name[len(prefix):]
                                    filtered_rows.append([original_name])
                            rows = filtered_rows
                        else:
                            # For regular databases, use the standard filtering
                            columns, rows = process_show_tables_result_for_schema(
                                columns, rows, database_name, current_user.id
                            )
                    elif is_show_databases_query(query):
                        # Filter SHOW DATABASES result to include allowed databases and teacher's schemas
                        columns, rows = filter_show_databases_result_for_user(columns, rows, current_user)
                    
                    result = {
                        'columns': columns,
                        'data': rows
                    }
                else:
                    # For non-SELECT queries (INSERT, UPDATE, etc.)
                    affected_rows = cursor.rowcount
                    result = {
                        'columns': ['Result'],
                        'data': [[f"{affected_rows} row(s) affected"]]
                    }
            
            conn.close()
            return jsonify(result)
            
        except pymysql.err.OperationalError as e:
            error_code, error_message = e.args
            if error_code == 1049:  # Unknown database
                return jsonify({'error': f'Database "{database_name}" does not exist'}), 404
            elif error_code == 1045:  # Access denied
                return jsonify({'error': 'Access denied. Check your database credentials.'}), 403
            clean_message = error_message.replace('\n', ' ')
            return jsonify({'error': f'Database Error: {clean_message}'}), 400
        except pymysql.err.ProgrammingError as e:
            error_code, error_message = e.args
            clean_message = error_message.replace('\n', ' ')
            return jsonify({'error': f'SQL Error: {clean_message}'}), 400
        except Exception as e:
            return jsonify({'error': f'Unexpected error: {str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'error': f'An unexpected error occurred: {str(e)}'}), 500

@teacher.route('/api/get-available-databases', methods=['GET'])
@login_required
@teacher_required
def get_available_databases():
    """API endpoint to get a list of available MySQL databases."""
    try:
        from app.models import AllowedDatabase
        from app.models.schema_import import SchemaImport
        
        available_databases = []
        
        # First check if there are any allowed databases configured by admin
        allowed_databases = AllowedDatabase.get_active_database_names()
        if allowed_databases:
            available_databases.extend(allowed_databases)
        
        # Add teacher's own imported schemas
        try:
            teacher_schemas = SchemaImport.query.filter_by(created_by=current_user.id).filter(
                SchemaImport.active_schema_name.isnot(None)
            ).all()
            
            # Add each schema's name as an available database option
            for schema in teacher_schemas:
                schema_name = schema.name
                if schema_name not in available_databases:
                    available_databases.append(schema_name)
        except Exception as e:
            logging.error(f"Error retrieving schemas for teacher {current_user.id}: {e}")

        # If we have databases to show (either allowed or teacher schemas), return them
        if available_databases:
            available_databases.sort()
            return jsonify({'databases': available_databases})
        
        # Fallback: if no allowed databases configured and no teacher schemas,
        # return all databases except system ones (backward compatibility)
        conn = pymysql.connect(
            host=os.getenv('MYSQL_HOST', ''),
            user=os.getenv('MYSQL_USER', ''),
            password=os.getenv('MYSQL_PASSWORD', ''),
            port=os.getenv('MYSQL_PORT', 3306),
            cursorclass=pymysql.cursors.DictCursor
        )
        
        # Execute query to get databases
        with conn.cursor() as cursor:
            cursor.execute("SHOW DATABASES")
            all_databases = [db['Database'] for db in cursor.fetchall()]
            
            # For teachers, we include all databases except a few system ones
            system_dbs = ['sys']  # Keep information_schema, mysql, performance_schema for admin purposes
            user_databases = [db for db in all_databases if db not in system_dbs]
            
            # Sort alphabetically
            user_databases.sort()
        
        conn.close()
        return jsonify({'databases': user_databases})
        
    except Exception as e:
        return jsonify({'error': str(e), 'databases': []}), 500
